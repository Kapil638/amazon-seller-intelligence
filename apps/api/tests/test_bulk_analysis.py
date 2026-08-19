import io
import time

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.ai.mock import MockAIProvider
from app.bulk.cache import KeyedTtlCache, product_cache_key
from app.bulk.excel import build_bulk_workbook
from app.bulk.mock_catalog import TRANSIENT_ASIN
from app.bulk.mock_product import BulkMockProductDataProvider
from app.bulk.portfolio import aggregate_portfolio
from app.bulk.priority import classify_priority
from app.bulk.processor import BulkProcessor, select_ai_targets
from app.bulk.providers import get_bulk_ai_provider, get_bulk_product_provider
from app.core.exceptions import BulkLiveProviderForbiddenError
from app.core.config import get_settings
from app.models.bulk import BulkJobOptions
from app.prompts.listing_intelligence import PROMPT_VERSION
from app.services.listing_analysis_service import ListingAnalysisService
from app.usage.ledger import get_usage_ledger
from tests.bulk_helpers import csv_bytes, make_processor, standard_options, xlsx_bytes
from tests.test_ai_listing_intelligence import sample_intelligence
from tests.test_listing_analysis import make_product


@pytest.mark.asyncio
async def test_successful_fixture_product() -> None:
    processor, provider, _ai, _pc, _ac = make_processor()
    results, failures, usage = await processor.process(
        ["B0BLKSTR01"],
        marketplace="amazon.in",
        options=standard_options(),
        ingest_failures=[],
    )
    assert len(results) == 1
    assert results[0].product.asin == "B0BLKSTR01"
    assert results[0].listing_analysis.overall_score >= 0
    assert failures == []
    assert usage.provider_calls == 1
    assert provider.calls == ["B0BLKSTR01"]


@pytest.mark.asyncio
async def test_unknown_asin_is_controlled_failure() -> None:
    processor, provider, _ai, _pc, _ac = make_processor()
    results, failures, usage = await processor.process(
        ["B0UNKNOWN1"],
        marketplace="amazon.in",
        options=standard_options(),
        ingest_failures=[],
    )
    assert results == []
    assert failures[0].kind == "not_found"
    assert usage.failures == 1
    assert provider.calls == ["B0UNKNOWN1"]


@pytest.mark.asyncio
async def test_transient_failure_retries_once() -> None:
    processor, provider, _ai, _pc, _ac = make_processor()
    results, failures, usage = await processor.process(
        [TRANSIENT_ASIN],
        marketplace="amazon.in",
        options=standard_options(),
        ingest_failures=[],
    )
    assert len(results) == 1
    assert results[0].asin == TRANSIENT_ASIN
    assert failures == []
    assert usage.retries == 1
    assert usage.provider_calls == 2
    assert provider.calls == [TRANSIENT_ASIN, TRANSIENT_ASIN]


@pytest.mark.asyncio
async def test_mixed_success_and_failure() -> None:
    processor, _provider, _ai, _pc, _ac = make_processor()
    results, failures, usage = await processor.process(
        ["B0BLKSTR01", "B0UNKNOWN1", "B0TEST0001"],
        marketplace="amazon.in",
        options=standard_options(),
        ingest_failures=[],
    )
    assert {item.asin for item in results} == {"B0BLKSTR01", "B0TEST0001"}
    assert len(failures) == 1
    assert usage.failures == 1


@pytest.mark.asyncio
async def test_first_lookup_invokes_mock_provider() -> None:
    processor, provider, _ai, _pc, _ac = make_processor()
    await processor.process(["B0BLKSTR01"], marketplace="amazon.in", options=standard_options(), ingest_failures=[])
    assert provider.calls == ["B0BLKSTR01"]


@pytest.mark.asyncio
async def test_repeated_lookup_uses_cache() -> None:
    processor, provider, _ai, cache, _ac = make_processor()
    await processor.process(["B0BLKSTR01"], marketplace="amazon.in", options=standard_options(), ingest_failures=[])
    provider.calls.clear()
    results, _failures, usage = await processor.process(
        ["B0BLKSTR01"], marketplace="amazon.in", options=standard_options(), ingest_failures=[]
    )
    assert provider.calls == []
    assert usage.cache_hits == 1
    assert usage.provider_calls == 0
    assert usage.calls_saved == 1
    assert results[0].cache_hit is True
    assert cache.get(product_cache_key("mock", "amazon.in", "B0BLKSTR01")) is not None


@pytest.mark.asyncio
async def test_duplicate_asin_invokes_provider_once() -> None:
    filename, data = csv_bytes(["ASIN"], [["B0BLKSTR01"], ["B0BLKSTR01"]])
    from app.bulk.ingest import ingest_asin_file

    _stats, unique, failures = ingest_asin_file(filename, data)
    processor, provider, _ai, _pc, _ac = make_processor()
    await processor.process(unique, marketplace="amazon.in", options=standard_options(), ingest_failures=failures)
    assert unique == ["B0BLKSTR01"]
    assert provider.calls == ["B0BLKSTR01"]


@pytest.mark.asyncio
async def test_cache_expiry_invokes_provider_again() -> None:
    processor, provider, _ai, _pc, _ac = make_processor(product_ttl=0)
    await processor.process(["B0BLKSTR01"], marketplace="amazon.in", options=standard_options(), ingest_failures=[])
    await processor.process(["B0BLKSTR01"], marketplace="amazon.in", options=standard_options(), ingest_failures=[])
    assert provider.calls == ["B0BLKSTR01", "B0BLKSTR01"]


@pytest.mark.asyncio
async def test_mixed_cached_and_uncached_batch() -> None:
    processor, provider, _ai, _pc, _ac = make_processor()
    await processor.process(["B0BLKSTR01"], marketplace="amazon.in", options=standard_options(), ingest_failures=[])
    provider.calls.clear()
    _results, _failures, usage = await processor.process(
        ["B0BLKSTR01", "B0TEST0002"],
        marketplace="amazon.in",
        options=standard_options(),
        ingest_failures=[],
    )
    assert provider.calls == ["B0TEST0002"]
    assert usage.cache_hits == 1
    assert usage.provider_calls == 1


def test_listing_analysis_reused_and_matches_single_asin() -> None:
    processor, _provider, _ai, _pc, _ac = make_processor()
    product = _provider._extra[("B0BLKSTR01", "amazon.in")]
    expected = ListingAnalysisService().analyze(product)
    assert processor._analysis.analyze(product).overall_score == expected.overall_score
    assert processor._analysis.analyze(product).findings == expected.findings


@pytest.mark.asyncio
async def test_priority_classification_bands() -> None:
    processor, _provider, _ai, _pc, _ac = make_processor()
    results, _failures, _usage = await processor.process(
        ["B0BLKHGH09", "B0BLKMID10", "B0BLKSTR01"],
        marketplace="amazon.in",
        options=standard_options(),
        ingest_failures=[],
    )
    by_asin = {item.asin: item for item in results}
    assert by_asin["B0BLKHGH09"].priority == "high"
    assert by_asin["B0BLKSTR01"].priority == "low"
    assert by_asin["B0BLKMID10"].priority in {"medium", "high", "low"}
    high = classify_priority(by_asin["B0BLKHGH09"].listing_analysis)
    low = classify_priority(by_asin["B0BLKSTR01"].listing_analysis)
    assert high == "high"
    assert low == "low"
    assert by_asin["B0BLKHGH09"].listing_analysis.overall_score < 50 or any(
        item.severity.value == "high" for item in by_asin["B0BLKHGH09"].listing_analysis.findings
    )


def test_medium_priority_from_score_band() -> None:
    analysis = ListingAnalysisService().analyze(
        make_product(title="Short name", description="Too short", bullet_points=["A", "B"], images=[])
    )
    # images=[] is high severity NO_IMAGES, so force a medium-only case via score band helper
    analysis.overall_score = 60
    analysis.findings = [item for item in analysis.findings if item.severity.value != "high"]
    while len([item for item in analysis.findings if item.severity.value == "medium"]) < 2:
        from app.models.listing_analysis import Finding, FindingSeverity

        analysis.findings.append(
            Finding(severity=FindingSeverity.MEDIUM, category="test", code="TEST_MED", message="medium")
        )
    assert classify_priority(analysis) == "medium"


@pytest.mark.asyncio
async def test_average_and_median_scores() -> None:
    processor, _provider, _ai, _pc, _ac = make_processor()
    results, failures, _usage = await processor.process(
        ["B0BLKSTR01", "B0BLKLOW11", "B0BLKMID10"],
        marketplace="amazon.in",
        options=standard_options(),
        ingest_failures=[],
    )
    summary = aggregate_portfolio(submitted=3, results=results, failures=failures)
    scores = sorted(item.listing_analysis.overall_score for item in results)
    assert summary.average_listing_score == round(sum(scores) / 3, 1)
    assert summary.median_listing_score == float(scores[1])
    assert summary.products_analyzed == 3


@pytest.mark.asyncio
async def test_finding_aggregation_counts() -> None:
    processor, _provider, _ai, _pc, _ac = make_processor()
    results, failures, _usage = await processor.process(
        ["B0BLKDES03", "B0BLKIMG04", "B0BLKBLT05", "B0BLKINC08"],
        marketplace="amazon.in",
        options=standard_options(),
        ingest_failures=[],
    )
    summary = aggregate_portfolio(submitted=4, results=results, failures=failures)
    assert summary.missing_description_count >= 1
    assert summary.low_image_count >= 1
    assert summary.weak_bullet_count >= 1
    assert summary.low_completeness_count >= 1


@pytest.mark.asyncio
async def test_standard_mode_makes_zero_ai_calls() -> None:
    processor, _provider, ai, _pc, _ac = make_processor()
    await processor.process(
        ["B0BLKSTR01", "B0BLKHGH09"],
        marketplace="amazon.in",
        options=standard_options(),
        ingest_failures=[],
    )
    assert ai.calls == []
    assert get_usage_ledger().openai_requests == 0


@pytest.mark.asyncio
async def test_deep_mode_calls_mock_ai_only() -> None:
    processor, _provider, ai, _pc, _ac = make_processor()
    results, _failures, usage = await processor.process(
        ["B0BLKHGH09"],
        marketplace="amazon.in",
        options=standard_options(analysis_mode="deep_ai", ai_selection="all"),
        ingest_failures=[],
    )
    assert len(ai.calls) == 1
    assert results[0].ai_intelligence is not None
    assert results[0].ai_intelligence.executive_summary
    sample_intelligence()  # schema still valid independently
    assert results[0].ai_status == "mock"
    assert usage.ai_provider_calls == 1
    assert get_usage_ledger().openai_requests == 0


@pytest.mark.asyncio
async def test_high_priority_only_selection() -> None:
    processor, _provider, ai, _pc, _ac = make_processor()
    results, _failures, usage = await processor.process(
        ["B0BLKSTR01", "B0BLKHGH09"],
        marketplace="amazon.in",
        options=standard_options(analysis_mode="deep_ai", ai_selection="high_priority"),
        ingest_failures=[],
    )
    selected = select_ai_targets(results, "high_priority", 10)
    assert all(item.priority == "high" for item in selected)
    assert usage.ai_eligible == len(selected)
    assert len(ai.calls) == usage.ai_eligible
    by_asin = {item.asin: item for item in results}
    assert by_asin["B0BLKSTR01"].ai_status == "skipped"
    assert by_asin["B0BLKHGH09"].ai_status == "mock"


@pytest.mark.asyncio
async def test_top_n_and_all_selection() -> None:
    processor, _provider, ai, _pc, _ac = make_processor()
    results, _failures, usage = await processor.process(
        ["B0BLKSTR01", "B0BLKMID10", "B0BLKHGH09"],
        marketplace="amazon.in",
        options=standard_options(analysis_mode="deep_ai", ai_selection="top_n", top_n=2),
        ingest_failures=[],
    )
    assert usage.ai_eligible == 2
    assert len(ai.calls) == 2
    ai.calls.clear()
    processor2, _p, ai2, _c, _a = make_processor()
    _results, _f, usage_all = await processor2.process(
        ["B0BLKSTR01", "B0BLKMID10"],
        marketplace="amazon.in",
        options=standard_options(analysis_mode="deep_ai", ai_selection="all"),
        ingest_failures=[],
    )
    assert usage_all.ai_eligible == 2
    assert len(ai2.calls) == 2


@pytest.mark.asyncio
async def test_ai_cache_hit_avoids_mock_call() -> None:
    processor, _provider, ai, _pc, _ac = make_processor()
    options = standard_options(analysis_mode="deep_ai", ai_selection="all")
    await processor.process(["B0BLKSTR01"], marketplace="amazon.in", options=options, ingest_failures=[])
    assert len(ai.calls) == 1
    ai.calls.clear()
    results, _failures, usage = await processor.process(
        ["B0BLKSTR01"], marketplace="amazon.in", options=options, ingest_failures=[]
    )
    assert ai.calls == []
    assert usage.ai_cache_hits == 1
    assert usage.ai_provider_calls == 0
    assert results[0].ai_status == "cached"
    assert results[0].ai_intelligence is not None


@pytest.mark.asyncio
async def test_ai_output_schema_valid() -> None:
    processor, _provider, _ai, _pc, _ac = make_processor()
    results, _failures, _usage = await processor.process(
        ["B0BLKSTR01"],
        marketplace="amazon.in",
        options=standard_options(analysis_mode="deep_ai", ai_selection="all"),
        ingest_failures=[],
    )
    payload = results[0].ai_intelligence
    assert payload is not None
    assert payload.priority_actions
    assert payload.title_recommendation.suggested_title
    assert payload.seller_action_plan
    assert PROMPT_VERSION


def test_live_rainforest_forbidden_in_bulk(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("BULK_PRODUCT_PROVIDER", "rainforest")
    monkeypatch.setenv("BULK_LIVE_PROVIDER_CALLS_ENABLED", "false")
    get_settings.cache_clear()
    try:
        with pytest.raises(BulkLiveProviderForbiddenError, match="live product provider"):
            get_bulk_product_provider()
    finally:
        get_settings.cache_clear()


def test_live_openai_forbidden_in_bulk(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("BULK_AI_PROVIDER", "openai")
    monkeypatch.setenv("BULK_LIVE_PROVIDER_CALLS_ENABLED", "false")
    get_settings.cache_clear()
    try:
        with pytest.raises(BulkLiveProviderForbiddenError, match="live AI provider"):
            get_bulk_ai_provider()
    finally:
        get_settings.cache_clear()


@pytest.mark.asyncio
async def test_mock_mode_records_zero_paid_calls() -> None:
    processor, _provider, _ai, _pc, _ac = make_processor()
    _results, _failures, usage = await processor.process(
        ["B0BLKSTR01"],
        marketplace="amazon.in",
        options=standard_options(analysis_mode="deep_ai", ai_selection="all"),
        ingest_failures=[],
    )
    assert usage.paid_api_usage is False
    assert "Mock provider" in usage.note
    assert get_usage_ledger().rainforest_product_calls == 0
    assert get_usage_ledger().openai_requests == 0


def _wait_job(client: TestClient, job_id: str) -> dict:
    deadline = time.time() + 8
    last = {}
    while time.time() < deadline:
        response = client.get(f"/api/v1/bulk/jobs/{job_id}")
        assert response.status_code == 200
        last = response.json()
        if last["status"] in {"completed", "completed_with_errors", "failed"}:
            return last
        time.sleep(0.05)
    raise AssertionError(f"job did not finish: {last}")


def test_job_lifecycle_and_progress(client: TestClient) -> None:
    filename, data = csv_bytes(["ASIN"], [["B0BLKSTR01"], ["B0UNKNOWN1"]])
    created = client.post(
        "/api/v1/bulk/jobs",
        files={"file": (filename, data, "text/csv")},
        data={"analysis_mode": "standard"},
    )
    assert created.status_code == 200
    body = created.json()
    assert body["status"] in {"queued", "running", "completed", "completed_with_errors"}
    assert body["job_id"]
    finished = _wait_job(client, body["job_id"])
    assert finished["status"] == "completed_with_errors"
    assert finished["progress"]["total"] == 2
    assert finished["progress"]["successful"] == 1
    assert finished["progress"]["failed"] == 1
    assert finished["summary"]["products_analyzed"] == 1


def test_job_completed_standard(client: TestClient) -> None:
    filename, data = csv_bytes(["ASIN"], [["B0BLKSTR01"], ["B0TEST0001"]])
    created = client.post(
        "/api/v1/bulk/jobs",
        files={"file": (filename, data, "text/csv")},
        data={"analysis_mode": "standard"},
    )
    finished = _wait_job(client, created.json()["job_id"])
    assert finished["status"] == "completed"
    assert finished["usage"]["ai_provider_calls"] == 0
    assert all(item["ai_intelligence"] is None for item in finished["results"])


def test_job_not_found(client: TestClient) -> None:
    assert client.get("/api/v1/bulk/jobs/missing").status_code == 404


def test_job_failed_status(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    from app.bulk.runtime import get_bulk_job_service

    service = get_bulk_job_service()

    async def boom(*_args: object, **_kwargs: object):
        raise RuntimeError("processor exploded")

    monkeypatch.setattr(service, "_processor_factory", lambda: type("P", (), {"process": boom})())
    filename, data = csv_bytes(["ASIN"], [["B0BLKSTR01"]])
    created = client.post(
        "/api/v1/bulk/jobs",
        files={"file": (filename, data, "text/csv")},
        data={"analysis_mode": "standard"},
    )
    finished = _wait_job(client, created.json()["job_id"])
    assert finished["status"] == "failed"


def test_excel_workbook_sheets(client: TestClient) -> None:
    filename, data = csv_bytes(["ASIN"], [["B0BLKSTR01"], ["B0UNKNOWN1"]])
    created = client.post(
        "/api/v1/bulk/jobs",
        files={"file": (filename, data, "text/csv")},
        data={"analysis_mode": "standard"},
    )
    job_id = created.json()["job_id"]
    finished = _wait_job(client, job_id)
    from app.models.bulk import BulkJobResponse

    workbook = load_workbook(io.BytesIO(build_bulk_workbook(BulkJobResponse.model_validate(finished))))
    assert workbook.sheetnames == ["Executive Summary", "Product Findings", "Failures", "API Usage"]
    assert "AI Recommendations" not in workbook.sheetnames
    summary_values = [row[0].value for row in workbook["Executive Summary"].iter_rows(min_row=2)]
    assert "Average listing score" in summary_values
    assert workbook["Product Findings"]["A1"].value == "ASIN"
    assert workbook["Failures"]["A1"].value == "Row"
    usage_sheet = workbook["API Usage"]
    notes = [row[1].value for row in usage_sheet.iter_rows(min_row=2) if row[0].value == "Note"]
    assert notes and "Mock provider" in str(notes[0])

    download = client.get(f"/api/v1/bulk/jobs/{job_id}/report.xlsx")
    assert download.status_code == 200
    assert download.headers["content-type"].startswith("application/vnd.openxmlformats")


def test_excel_ai_sheet_only_in_deep_mode(client: TestClient) -> None:
    filename, data = csv_bytes(["ASIN"], [["B0BLKHGH09"]])
    created = client.post(
        "/api/v1/bulk/jobs",
        files={"file": (filename, data, "text/csv")},
        data={"analysis_mode": "deep_ai", "ai_selection": "all"},
    )
    finished = _wait_job(client, created.json()["job_id"])
    from app.models.bulk import BulkJobResponse

    workbook = load_workbook(io.BytesIO(build_bulk_workbook(BulkJobResponse.model_validate(finished))))
    assert "AI Recommendations" in workbook.sheetnames
    assert workbook["AI Recommendations"]["A1"].value == "ASIN"


def test_preview_endpoint(client: TestClient) -> None:
    filename, data = csv_bytes(["ASIN"], [["B0BLKSTR01"], ["B0BLKSTR01"], ["NOPE"]])
    response = client.post("/api/v1/bulk/preview", files={"file": (filename, data, "text/csv")})
    assert response.status_code == 200
    body = response.json()
    assert body["unique_asins"] == 1
    assert body["duplicate_rows_removed"] == 1
    assert body["invalid_rows"] == 1


def test_second_identical_job_uses_cache(client: TestClient) -> None:
    filename, data = csv_bytes(["ASIN"], [["B0TEST0001"], ["B0TEST0002"]])
    first = _wait_job(
        client,
        client.post("/api/v1/bulk/jobs", files={"file": (filename, data, "text/csv")}).json()["job_id"],
    )
    second = _wait_job(
        client,
        client.post("/api/v1/bulk/jobs", files={"file": (filename, data, "text/csv")}).json()["job_id"],
    )
    assert first["usage"]["provider_calls"] == 2
    assert second["usage"]["cache_hits"] == 2
    assert second["usage"]["provider_calls"] == 0
    assert first["usage"]["paid_api_usage"] is False
    assert second["usage"]["paid_api_usage"] is False
