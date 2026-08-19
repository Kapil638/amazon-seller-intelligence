"""Load CSV/XLSX into a header + row table. Cell values only; no formulas."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from zipfile import BadZipFile

from openpyxl import load_workbook

from app.core.exceptions import ReportParseError, ReportUploadError
from app.reports.columns import ALL_KNOWN_NORMALIZED, normalize_header

SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
HEADER_SCAN_ROWS = 25


@dataclass(frozen=True)
class TabularFile:
    headers: list[str]
    rows: list[list[str]]
    source_format: str


def load_tabular_file(filename: str, data: bytes) -> TabularFile:
    extension = _extension(filename)
    if extension == ".csv":
        rows = _read_csv(data)
        source_format = "csv"
    elif extension == ".xlsx":
        rows = _read_xlsx(data)
        source_format = "xlsx"
    else:
        raise ReportUploadError("Upload a .csv or .xlsx file.")

    if not rows:
        raise ReportUploadError("This file is empty.")

    header_index = _find_header_row(rows)
    if header_index is None:
        raise ReportParseError("No recognizable Amazon report headers were found.")

    headers = [_cell(value) for value in rows[header_index]]
    body: list[list[str]] = []
    width = len(headers)
    for row in rows[header_index + 1 :]:
        cells = [_cell(value) for value in row]
        if width > len(cells):
            cells.extend([""] * (width - len(cells)))
        elif len(cells) > width:
            cells = cells[:width]
        if all(not cell.strip() for cell in cells):
            continue
        body.append(cells)
    return TabularFile(headers=headers, rows=body, source_format=source_format)


def _extension(filename: str) -> str:
    name = filename.rsplit("/", 1)[-1].rsplit("\\", 1)[-1].strip().casefold()
    if "." not in name:
        return ""
    return "." + name.rsplit(".", 1)[-1]


def _read_csv(data: bytes) -> list[list[str]]:
    text = _decode_bytes(data)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    return [list(row) for row in reader]


def _decode_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "utf-16-le", "cp1252", "latin-1"):
        try:
            text = data.decode(encoding)
        except UnicodeDecodeError:
            continue
        if text.strip():
            return text
    raise ReportParseError("This CSV file could not be read.")


def _read_xlsx(data: bytes) -> list[list[str]]:
    buffer = io.BytesIO(data)
    try:
        workbook = load_workbook(
            buffer,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except BadZipFile as exc:
        raise ReportParseError("This spreadsheet is corrupted or is not a valid .xlsx file.") from exc
    except Exception as exc:
        message = str(exc).casefold()
        if "password" in message:
            raise ReportParseError("Password-protected spreadsheets are not supported.") from exc
        raise ReportParseError("This spreadsheet could not be read.") from exc

    try:
        sheet = workbook.active
        if sheet is None:
            raise ReportParseError("This spreadsheet has no worksheet.")
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([_cell(value) for value in row])
        return rows
    finally:
        workbook.close()


def _find_header_row(rows: list[list[str]]) -> int | None:
    best_index: int | None = None
    best_score = 0
    scan = min(len(rows), HEADER_SCAN_ROWS)
    for index in range(scan):
        score = 0
        for cell in rows[index]:
            key = normalize_header(_cell(cell))
            if key in ALL_KNOWN_NORMALIZED:
                score += 1
        if score > best_score:
            best_score = score
            best_index = index
    if best_score >= 3:
        return best_index
    return None


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
