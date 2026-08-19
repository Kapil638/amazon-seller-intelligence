from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from app.models.bulk import BulkJobResponse

_HEADER_FONT = Font(bold=True)


def build_bulk_workbook(job: BulkJobResponse) -> bytes:
    workbook = Workbook()
    _executive_summary(workbook, job)
    _product_findings(workbook, job)
    _failures(workbook, job)
    _api_usage(workbook, job)
    if job.options.analysis_mode == "deep_ai":
        _ai_recommendations(workbook, job)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _sheet(workbook: Workbook, title: str, index: int):
    if index == 0:
        sheet = workbook.active
        sheet.title = title
        return sheet
    return workbook.create_sheet(title)


def _write_header(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = _HEADER_FONT


def _executive_summary(workbook: Workbook, job: BulkJobResponse) -> None:
    sheet = _sheet(workbook, "Executive Summary", 0)
    summary = job.summary
    rows = [
        ("Job ID", job.job_id),
        ("Status", job.status),
        ("Analysis mode", job.options.analysis_mode),
        ("Filename", job.ingest.filename),
        ("Products submitted", summary.products_submitted if summary else job.ingest.unique_asins),
        ("Products analyzed", summary.products_analyzed if summary else 0),
        ("Products failed", summary.products_failed if summary else 0),
        ("Average listing score", summary.average_listing_score if summary else None),
        ("Median listing score", summary.median_listing_score if summary else None),
        ("High priority", summary.high_priority_count if summary else 0),
        ("Medium priority", summary.medium_priority_count if summary else 0),
        ("Low priority", summary.low_priority_count if summary else 0),
        ("Missing description", summary.missing_description_count if summary else 0),
        ("Low image count", summary.low_image_count if summary else 0),
        ("Weak bullets", summary.weak_bullet_count if summary else 0),
        ("Low completeness", summary.low_completeness_count if summary else 0),
        ("Invalid input rows", job.ingest.invalid_rows),
        ("Duplicates removed", job.ingest.duplicate_rows_removed),
    ]
    _write_header(sheet, ["Metric", "Value"])
    for metric, value in rows:
        sheet.append([metric, value])


def _product_findings(workbook: Workbook, job: BulkJobResponse) -> None:
    sheet = _sheet(workbook, "Product Findings", 1)
    _write_header(
        sheet,
        [
            "ASIN",
            "Title",
            "Brand",
            "Price",
            "Rating",
            "Reviews",
            "Images",
            "Overall Score",
            "Title Score",
            "Bullet Score",
            "Description Score",
            "Image Score",
            "Completeness Score",
            "Social Proof Score",
            "Priority",
            "Finding 1",
            "Finding 2",
            "Finding 3",
        ],
    )
    for item in job.results:
        product = item.product
        analysis = item.listing_analysis
        findings = [finding.message for finding in analysis.findings[:3]]
        while len(findings) < 3:
            findings.append("")
        price = None
        if product.price is not None:
            price = f"{product.price.amount} {product.price.currency}"
        sheet.append(
            [
                item.asin,
                product.title,
                product.brand,
                price,
                product.rating,
                product.review_count,
                len(product.images),
                analysis.overall_score,
                analysis.sections.title.score,
                analysis.sections.bullets.score,
                analysis.sections.description.score,
                analysis.sections.images.score,
                analysis.sections.completeness.score,
                analysis.sections.social_proof.score,
                item.priority,
                *findings,
            ]
        )


def _failures(workbook: Workbook, job: BulkJobResponse) -> None:
    sheet = _sheet(workbook, "Failures", 2)
    _write_header(sheet, ["Row", "Input ASIN", "Reason"])
    for item in job.failures:
        sheet.append([item.row, item.input_asin, item.reason])


def _api_usage(workbook: Workbook, job: BulkJobResponse) -> None:
    sheet = _sheet(workbook, "API Usage", 3)
    usage = job.usage
    _write_header(sheet, ["Metric", "Value"])
    rows = [
        ("Provider mode", usage.product_provider),
        ("Paid API usage", usage.paid_api_usage),
        ("Note", usage.note),
        ("Requested ASINs", usage.requested_asins),
        ("Cache hits", usage.cache_hits),
        ("Provider calls", usage.provider_calls),
        ("Calls saved", usage.calls_saved),
        ("Failures", usage.failures),
        ("Retries", usage.retries),
        ("AI provider", usage.ai_provider),
        ("AI eligible", usage.ai_eligible),
        ("AI cache hits", usage.ai_cache_hits),
        ("AI provider calls", usage.ai_provider_calls),
        ("AI calls saved", usage.ai_calls_saved),
    ]
    for metric, value in rows:
        sheet.append([metric, value])


def _ai_recommendations(workbook: Workbook, job: BulkJobResponse) -> None:
    sheet = workbook.create_sheet("AI Recommendations")
    _write_header(
        sheet,
        ["ASIN", "Executive Summary", "Top Priority", "Suggested Title", "Top Action"],
    )
    for item in job.results:
        intel = item.ai_intelligence
        if intel is None:
            continue
        top_priority = intel.priority_actions[0].title if intel.priority_actions else ""
        top_action = intel.seller_action_plan[0].action if intel.seller_action_plan else ""
        sheet.append(
            [
                item.asin,
                intel.executive_summary,
                top_priority,
                intel.title_recommendation.suggested_title,
                top_action,
            ]
        )
